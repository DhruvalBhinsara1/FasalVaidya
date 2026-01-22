"""
Report Export Module
====================
Generates PDF, Excel, and CSV exports for crop health reports.
Supports single and bulk export with performance optimizations.
"""

import csv
import io
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
import logging

logger = logging.getLogger('fasalvaidya.export')

# ============================================
# CSV EXPORT
# ============================================

def export_to_csv(reports: List[Dict], include_headers: bool = True) -> str:
    """
    Export report data to CSV format.
    
    Args:
        reports: List of report data dictionaries
        include_headers: Whether to include header row
        
    Returns:
        CSV string
    """
    if not reports:
        return ""
    
    output = io.StringIO()
    
    # Define CSV columns
    fieldnames = [
        'scan_id', 'scan_date', 'crop_name', 'crop_name_hi',
        'n_score', 'n_severity', 'p_score', 'p_severity', 
        'k_score', 'k_severity', 'mg_score', 'mg_severity',
        'overall_score', 'health_status', 'health_status_hi',
        'recommended_rescan_date', 'critical_nutrients', 'attention_nutrients',
        'farmer_name', 'location'
    ]
    
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
    
    if include_headers:
        writer.writeheader()
    
    for report in reports:
        current_scan = report.get('current_scan', {})
        health = report.get('health_classification', {})
        recs = report.get('recommendations', {})
        field = report.get('field_info', {})
        farmer = report.get('farmer_info', {})
        nutrients = current_scan.get('nutrients', {})
        
        row = {
            'scan_id': current_scan.get('scan_id', ''),
            'scan_date': current_scan.get('scan_date', ''),
            'crop_name': field.get('crop_name', ''),
            'crop_name_hi': field.get('crop_name_hi', ''),
            'n_score': nutrients.get('nitrogen', {}).get('score', ''),
            'n_severity': nutrients.get('nitrogen', {}).get('severity', ''),
            'p_score': nutrients.get('phosphorus', {}).get('score', ''),
            'p_severity': nutrients.get('phosphorus', {}).get('severity', ''),
            'k_score': nutrients.get('potassium', {}).get('score', ''),
            'k_severity': nutrients.get('potassium', {}).get('severity', ''),
            'mg_score': nutrients.get('magnesium', {}).get('score', '') if 'magnesium' in nutrients else '',
            'mg_severity': nutrients.get('magnesium', {}).get('severity', '') if 'magnesium' in nutrients else '',
            'overall_score': health.get('overall_score', ''),
            'health_status': health.get('label', ''),
            'health_status_hi': health.get('label_hi', ''),
            'recommended_rescan_date': recs.get('rescan', {}).get('recommended_date', ''),
            'critical_nutrients': ','.join(recs.get('summary', {}).get('critical_nutrients', [])),
            'attention_nutrients': ','.join(recs.get('summary', {}).get('attention_nutrients', [])),
            'farmer_name': farmer.get('name', ''),
            'location': farmer.get('location', '')
        }
        
        writer.writerow(row)
    
    return output.getvalue()


# ============================================
# EXCEL EXPORT
# ============================================

def export_to_excel(reports: List[Dict]) -> bytes:
    """
    Export report data to Excel format.
    
    Args:
        reports: List of report data dictionaries
        
    Returns:
        Excel file bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Crop Health Report"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4C763B", end_color="4C763B", fill_type="solid")
    healthy_fill = PatternFill(start_color="DBFFCB", end_color="DBFFCB", fill_type="solid")
    attention_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    critical_fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Scan ID', 'Date', 'Crop', 'N Score', 'N Status', 
        'P Score', 'P Status', 'K Score', 'K Status',
        'Overall Score', 'Health Status', 'Next Scan', 'Issues'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data rows
    for row_idx, report in enumerate(reports, 2):
        current_scan = report.get('current_scan', {})
        health = report.get('health_classification', {})
        recs = report.get('recommendations', {})
        field = report.get('field_info', {})
        nutrients = current_scan.get('nutrients', {})
        
        data = [
            current_scan.get('scan_id', ''),
            current_scan.get('scan_date', '')[:10] if current_scan.get('scan_date') else '',
            field.get('crop_name', ''),
            nutrients.get('nitrogen', {}).get('health_score', ''),
            nutrients.get('nitrogen', {}).get('severity', ''),
            nutrients.get('phosphorus', {}).get('health_score', ''),
            nutrients.get('phosphorus', {}).get('severity', ''),
            nutrients.get('potassium', {}).get('health_score', ''),
            nutrients.get('potassium', {}).get('severity', ''),
            health.get('overall_score', ''),
            health.get('label', ''),
            recs.get('rescan', {}).get('recommended_date', ''),
            recs.get('summary', {}).get('total_issues', 0)
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Color code health status
        status = health.get('status', '')
        if status == 'good':
            ws.cell(row=row_idx, column=11).fill = healthy_fill
        elif status == 'average':
            ws.cell(row=row_idx, column=11).fill = attention_fill
        elif status == 'unhealthy':
            ws.cell(row=row_idx, column=11).fill = critical_fill
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


# ============================================
# PDF EXPORT
# ============================================

def export_to_pdf(report: Dict) -> bytes:
    """
    Export single report to PDF format.
    
    Args:
        report: Report data dictionary
        
    Returns:
        PDF file bytes
    """
    try:
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        logger.error("reportlab not installed. Install with: pip install reportlab")
        raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=rl_colors.HexColor('#4C763B'),
        spaceAfter=12
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=rl_colors.HexColor('#043915'),
        spaceAfter=8,
        spaceBefore=16
    )
    
    elements = []
    
    # Title
    elements.append(Paragraph("🌿 FasalVaidya - Crop Health Report", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Report metadata
    generated_at = report.get('generated_at', datetime.now().isoformat())
    elements.append(Paragraph(f"Generated: {generated_at[:10]}", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Section 1: Farmer & Field Summary
    elements.append(Paragraph("1. Farmer & Field Summary", heading_style))
    field = report.get('field_info', {})
    farmer = report.get('farmer_info', {})
    
    summary_data = [
        ['Farmer Name:', farmer.get('name', 'Guest User')],
        ['Location:', farmer.get('location', 'Not specified')],
        ['Crop:', f"{field.get('crop_icon', '')} {field.get('crop_name', '')}"],
        ['Season:', field.get('season', 'Not specified')]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Section 2: Current Scan Metrics
    elements.append(Paragraph("2. Current Scan Metrics", heading_style))
    current = report.get('current_scan', {})
    nutrients = current.get('nutrients', {})
    
    nutrient_data = [
        ['Nutrient', 'Health Score', 'Status', 'Confidence']
    ]
    
    for name, key in [('Nitrogen (N)', 'nitrogen'), ('Phosphorus (P)', 'phosphorus'), ('Potassium (K)', 'potassium')]:
        n = nutrients.get(key, {})
        nutrient_data.append([
            name,
            f"{n.get('health_score', 0):.1f}%",
            n.get('severity', 'healthy').title(),
            f"{n.get('confidence', 0):.1f}%"
        ])
    
    if 'magnesium' in nutrients:
        mg = nutrients['magnesium']
        nutrient_data.append([
            'Magnesium (Mg)',
            f"{mg.get('health_score', 0):.1f}%",
            mg.get('severity', 'healthy').title(),
            f"{mg.get('confidence', 0):.1f}%"
        ])
    
    nutrient_table = Table(nutrient_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    nutrient_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#4C763B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(nutrient_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Section 3: Health Status Classification
    elements.append(Paragraph("3. Health Status Classification", heading_style))
    health = report.get('health_classification', {})
    
    status_color = {
        'good': rl_colors.HexColor('#4C763B'),
        'average': rl_colors.HexColor('#FA8112'),
        'unhealthy': rl_colors.HexColor('#FF6363')
    }.get(health.get('status', 'average'), rl_colors.grey)
    
    status_data = [
        ['Overall Health Score:', f"{health.get('overall_score', 0):.1f}/100"],
        ['Status:', health.get('label', 'Unknown')],
        ['Severity Level:', health.get('severity', 'Unknown').title()]
    ]
    
    status_table = Table(status_data, colWidths=[2.5*inch, 3.5*inch])
    status_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (1, 1), (1, 1), status_color),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(status_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Section 4: Recommendations
    elements.append(Paragraph("4. Recommendations", heading_style))
    recs = report.get('recommendations', {})
    rescan = recs.get('rescan', {})
    fertilizer = recs.get('fertilizer', [])
    
    elements.append(Paragraph(f"<b>Next Scan Recommended:</b> {rescan.get('recommended_date', 'N/A')} ({rescan.get('interval_label', '')})", styles['Normal']))
    elements.append(Spacer(1, 0.1*inch))
    
    if fertilizer:
        elements.append(Paragraph("<b>Fertilizer Actions:</b>", styles['Normal']))
        for rec in fertilizer:
            priority_marker = "🔴" if rec.get('priority') == 'high' else "🟠" if rec.get('priority') == 'medium' else "🟢"
            elements.append(Paragraph(f"  {priority_marker} {rec.get('action_label', '')}", styles['Normal']))
    else:
        elements.append(Paragraph("✅ No fertilizer action needed. Crop is healthy!", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    output.seek(0)
    
    return output.getvalue()


def export_bulk_to_pdf(reports: List[Dict]) -> bytes:
    """
    Export multiple reports to a single PDF.
    
    Args:
        reports: List of report data dictionaries
        
    Returns:
        PDF file bytes
    """
    # For bulk, we create a summary PDF
    try:
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    except ImportError:
        raise ImportError("reportlab is required for PDF export")
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    elements.append(Paragraph("🌿 FasalVaidya - Bulk Crop Health Report", styles['Title']))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Paragraph(f"Total Reports: {len(reports)}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary table
    table_data = [['#', 'Scan ID', 'Date', 'Crop', 'N', 'P', 'K', 'Overall', 'Status', 'Next Scan']]
    
    for idx, report in enumerate(reports, 1):
        current = report.get('current_scan', {})
        health = report.get('health_classification', {})
        field = report.get('field_info', {})
        recs = report.get('recommendations', {})
        nutrients = current.get('nutrients', {})
        
        table_data.append([
            str(idx),
            str(current.get('scan_id', ''))[:8],
            str(current.get('scan_date', ''))[:10],
            field.get('crop_name', ''),
            f"{nutrients.get('nitrogen', {}).get('health_score', 0):.0f}",
            f"{nutrients.get('phosphorus', {}).get('health_score', 0):.0f}",
            f"{nutrients.get('potassium', {}).get('health_score', 0):.0f}",
            f"{health.get('overall_score', 0):.0f}",
            health.get('label', ''),
            recs.get('rescan', {}).get('recommended_date', '')
        ])
    
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#4C763B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#F4F8F4')]),
    ]))
    elements.append(table)
    
    doc.build(elements)
    output.seek(0)
    
    return output.getvalue()
